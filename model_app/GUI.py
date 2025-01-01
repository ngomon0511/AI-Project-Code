import sys
import os
import joblib
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog, QVBoxLayout
from PyQt5 import uic
from datetime import datetime
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment
from darts import TimeSeries  
from darts.models import RNNModel

import warnings
warnings.filterwarnings("ignore")
import logging
logging.disable(logging.CRITICAL)
logging.getLogger("darts").setLevel(logging.CRITICAL)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('model_app.ui', self)  

        self.predict_button.clicked.connect(self.predict_consumption)  
        self.export_button.clicked.connect(self.export_data)  
        self.move(190, 25)
        self.setFixedSize(1000, 661)            
        self.plotWidget.setLayout(QVBoxLayout())
        self.canvas = FigureCanvas(Figure(figsize=(12, 5)))
        self.plotWidget.layout().addWidget(self.canvas) 
        self.last_end_time = None
        self.backup_ts = None
        self.time = None
        self.demand = None

        self.load_data('past_data.csv')
        self.trained_model = RNNModel.load("saved_model/trained_model.darts")
        self.data_scaler = joblib.load("saved_model/data_scaler.pkl")
        self.covariates_scaler = joblib.load("saved_model/covariates_scaler.pkl")

    def load_data(self, file_path):
        df = pd.read_csv(file_path)
        df['Date'] = pd.to_datetime(df['Date'], format='%Y-%m-%d')
        df.set_index('Date', inplace=True)
        self.ts = TimeSeries.from_dataframe(df, value_cols='Demand')
        self.backup_ts = self.ts
        self.plot_data()

    def plot_data(self):
        self.canvas.figure.clf()
        ax = self.canvas.figure.add_subplot(111)
        self.time = self.ts.time_index
        self.demand = self.ts.values().flatten()  
        ax.plot(self.time, self.demand, color='blue', marker='o', markersize=4)  
        ax.set_title('Electricity Demand', fontsize=16)
        ax.set_xlabel('Time', fontsize=14)
        ax.set_ylabel('Demand (TWh)', fontsize=14) 
        self.canvas.draw()

    def predict_consumption(self):
        date_input = self.date_input.text()
        if not date_input:
            QMessageBox.critical(self, "Error", "Time input cannot be empty. Please enter a valid date!")
            return
        try:
            prediction_date = datetime.strptime(date_input, "%m/%Y")
        except ValueError:
            QMessageBox.critical(self, "Error", "Invalid Time format. Please use MM/YYYY!")
            return
        end_time = str(prediction_date)[:7]
        self.last_end_time = end_time

        if end_time != self.ts.time_index[-1].strftime('%Y-%m'):
            past_end_time = '2023-12'
            past_date = datetime.strptime(past_end_time, "%Y-%m")
            end_date = datetime.strptime(end_time, "%Y-%m")
            month_interval = (end_date.year - past_date.year) * 12 + (end_date.month - past_date.month)
            
            pred_data = self.trained_model.predict(n=month_interval, future_covariates=self.covariates_scaler)
            pred_data = [self.data_scaler.inverse_transform(p) for p in pred_data] if isinstance(pred_data, list) else self.data_scaler.inverse_transform(pred_data)
            pred_data_list = [p.values() for p in pred_data] if isinstance(pred_data, list) else pred_data.values()
            
            self.ts = self.backup_ts
            freq = self.ts.time_index.inferred_freq 
            for i in range(month_interval):
                next_month = self.ts.time_index[-1] + pd.DateOffset(months=1)
                next_month_index = pd.DatetimeIndex([next_month])
                val = pred_data_list[i]
                new_ts = TimeSeries.from_times_and_values(next_month_index, val, fill_missing_dates=True, freq=freq)
                self.ts = self.ts.append(new_ts)
            self.plot_data()

    def export_data(self):
        default_folder = os.path.join(os.getcwd(), 'output_data')
        if not os.path.exists(default_folder):
            os.makedirs(default_folder)
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Save File", default_folder, "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_name:
            if not file_name.endswith('.xlsx'):
                file_name += '.xlsx'
            data = self.ts.pd_dataframe()
            data.reset_index(inplace=True)  
            data.columns = ['Time', 'Demand (TWh)']  
            data['Time'] = data['Time'].dt.strftime('%Y-%m')

            wb = Workbook()
            ws = wb.active
            ws.append(['Time', 'Demand (TWh)'])
            for row in data.itertuples(index=False):
                ws.append(row)  
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
            col_letter = get_column_letter(ws.max_column)  
            ws.column_dimensions[col_letter].width = 14 
            wb.save(file_name)
            QMessageBox.information(self, "Success", f"Data successfully exported to {file_name}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
